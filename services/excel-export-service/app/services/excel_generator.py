"""
Excel Generation Service.

Handles Excel file generation with formatting, charts, and advanced features.
"""

import asyncio
import base64
import io
import logging
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import uuid

import httpx
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart, BarChart, ScatterChart, PieChart, AreaChart, RadarChart, BubbleChart
)
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from openpyxl.styles.numbers import FORMAT_GENERAL, FORMAT_NUMBER, FORMAT_DATE_YYYYMMDD2
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
import pandas as pd

from app.core.config import settings
from app.core.database import execute_query
from app.models.excel_models import (
    JobStatus, ExcelFormat, ChartType, CellDataType, Theme,
    WorkbookConfig, WorksheetConfig, ChartConfig, CellStyle, CellData,
    DataValidationRule
)


logger = logging.getLogger(__name__)


class ExcelGenerationError(Exception):
    """Excel generation error."""
    pass


class ExcelGenerator:
    """Excel generator service."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.active_jobs = {}
        self.theme_configs = self._load_theme_configs()
        self.chart_type_mapping = {
            ChartType.LINE: LineChart,
            ChartType.BAR: BarChart,
            ChartType.COLUMN: BarChart,
            ChartType.PIE: PieChart,
            ChartType.SCATTER: ScatterChart,
            ChartType.AREA: AreaChart,
            ChartType.RADAR: RadarChart,
            ChartType.BUBBLE: BubbleChart
        }
    
    def _load_theme_configs(self) -> Dict[str, Dict[str, Any]]:
        """Load theme configurations."""
        return {
            Theme.OFFICE: {
                "primary_color": "1F497D",
                "secondary_color": "4F81BD",
                "accent_color": "F79646",
                "background_color": "FFFFFF",
                "text_color": "000000",
                "header_color": "D9E2F3"
            },
            Theme.MODERN: {
                "primary_color": "2E75B6",
                "secondary_color": "70AD47",
                "accent_color": "FFC000",
                "background_color": "FFFFFF",
                "text_color": "44546A",
                "header_color": "E7E6E6"
            },
            Theme.COLORFUL: {
                "primary_color": "FF6B6B",
                "secondary_color": "4ECDC4",
                "accent_color": "FFE66D",
                "background_color": "FFFFFF",
                "text_color": "2C3E50",
                "header_color": "F8F9FA"
            },
            Theme.DARK: {
                "primary_color": "0078D4",
                "secondary_color": "107C10",
                "accent_color": "FF8C00",
                "background_color": "1F1F1F",
                "text_color": "FFFFFF",
                "header_color": "323130"
            },
            Theme.LIGHT: {
                "primary_color": "605E5C",
                "secondary_color": "0078D4",
                "accent_color": "D83B01",
                "background_color": "FFFFFF",
                "text_color": "323130",
                "header_color": "F3F2F1"
            }
        }
    
    async def generate_excel(
        self, 
        job_id: int, 
        workbook_config: WorkbookConfig,
        data_source: Dict[str, Any],
        output_format: ExcelFormat,
        theme: Theme,
        validation_rules: Optional[List[DataValidationRule]] = None
    ) -> Dict[str, Any]:
        """
        Generate Excel file from configuration.
        
        Args:
            job_id: Job ID
            workbook_config: Workbook configuration
            data_source: Data source configuration
            output_format: Output format
            theme: Theme
            validation_rules: Data validation rules
            
        Returns:
            Dictionary with file information
        """
        start_time = time.time()
        
        try:
            # Update job status
            await self._update_job_status(job_id, JobStatus.PROCESSING)
            
            # Add job to active jobs
            self.active_jobs[f"job:{job_id}"] = {
                "job_id": job_id,
                "start_time": start_time,
                "status": JobStatus.PROCESSING
            }
            
            # Create workbook
            workbook = Workbook()
            
            # Apply theme
            self._apply_theme(workbook, theme)
            
            # Set workbook properties
            if workbook_config.properties:
                self._set_workbook_properties(workbook, workbook_config.properties)
            
            # Remove default worksheet
            if workbook.active:
                workbook.remove(workbook.active)
            
            # Create worksheets
            for worksheet_config in workbook_config.worksheets:
                await self._create_worksheet(
                    workbook, worksheet_config, data_source, theme
                )
            
            # Apply data validation rules
            if validation_rules:
                self._apply_validation_rules(workbook, validation_rules)
            
            # Apply workbook protection
            if workbook_config.protection:
                self._apply_workbook_protection(workbook, workbook_config.protection)
            
            # Generate file
            file_info = await self._save_workbook(
                job_id, workbook, output_format, workbook_config.name
            )
            
            # Calculate generation time
            generation_time = int((time.time() - start_time) * 1000)
            
            # Update job completion
            await self._update_job_completion(job_id, file_info, generation_time)
            
            # Remove from active jobs
            self.active_jobs.pop(f"job:{job_id}", None)
            
            return file_info
            
        except Exception as e:
            logger.error(f"Excel generation failed for job {job_id}: {e}")
            
            # Calculate generation time
            generation_time = int((time.time() - start_time) * 1000)
            
            # Update job error
            await self._update_job_error(job_id, str(e), generation_time)
            
            # Remove from active jobs
            self.active_jobs.pop(f"job:{job_id}", None)
            
            raise ExcelGenerationError(f"Excel generation failed: {e}")
    
    async def _create_worksheet(
        self, 
        workbook: Workbook, 
        worksheet_config: WorksheetConfig,
        data_source: Dict[str, Any],
        theme: Theme
    ) -> None:
        """Create worksheet with data and formatting."""
        # Create worksheet
        worksheet = workbook.create_sheet(title=worksheet_config.name)
        
        # Add headers
        if worksheet_config.headers:
            for col, header in enumerate(worksheet_config.headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                if worksheet_config.header_style:
                    self._apply_cell_style(cell, worksheet_config.header_style, theme)
                else:
                    # Apply default header style
                    self._apply_default_header_style(cell, theme)
        
        # Add data
        start_row = 2 if worksheet_config.headers else 1
        for row_idx, row_data in enumerate(worksheet_config.data, start_row):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                self._populate_cell(cell, cell_data, theme)
        
        # Apply column widths
        if worksheet_config.column_widths:
            for col, width in worksheet_config.column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        # Apply row heights
        if worksheet_config.row_heights:
            for row, height in worksheet_config.row_heights.items():
                worksheet.row_dimensions[int(row)].height = height
        
        # Apply auto filter
        if worksheet_config.auto_filter and worksheet_config.headers:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Apply freeze panes
        if worksheet_config.freeze_panes:
            row = worksheet_config.freeze_panes.get("row", 1)
            col = worksheet_config.freeze_panes.get("col", 1)
            worksheet.freeze_panes = worksheet.cell(row=row, column=col)
        
        # Add charts
        if worksheet_config.charts:
            for chart_config in worksheet_config.charts:
                await self._add_chart(worksheet, chart_config, data_source, theme)
        
        # Apply worksheet protection
        if worksheet_config.protection:
            self._apply_worksheet_protection(worksheet, worksheet_config.protection)
    
    def _populate_cell(self, cell, cell_data: CellData, theme: Theme) -> None:
        """Populate cell with data and styling."""
        # Set value
        if cell_data.formula:
            cell.value = f"={cell_data.formula}"
        else:
            cell.value = cell_data.value
        
        # Apply data type formatting
        if cell_data.data_type == CellDataType.NUMBER:
            cell.number_format = FORMAT_NUMBER
        elif cell_data.data_type == CellDataType.DATE:
            cell.number_format = FORMAT_DATE_YYYYMMDD2
        elif cell_data.data_type == CellDataType.CURRENCY:
            cell.number_format = '"$"#,##0.00'
        elif cell_data.data_type == CellDataType.PERCENTAGE:
            cell.number_format = '0.00%'
        
        # Apply style
        if cell_data.style:
            self._apply_cell_style(cell, cell_data.style, theme)
        
        # Add comment
        if cell_data.comment:
            cell.comment = cell_data.comment
    
    def _apply_cell_style(self, cell, style: CellStyle, theme: Theme) -> None:
        """Apply cell style."""
        theme_config = self.theme_configs[theme]
        
        # Font
        font_color = style.font_color or theme_config["text_color"]
        cell.font = Font(
            name=style.font_name or "Arial",
            size=style.font_size or 11,
            bold=style.font_bold or False,
            italic=style.font_italic or False,
            color=font_color
        )
        
        # Fill
        if style.background_color:
            cell.fill = PatternFill(
                start_color=style.background_color,
                end_color=style.background_color,
                fill_type="solid"
            )
        
        # Border
        if style.border_style:
            border_color = style.border_color or "000000"
            border = Border(
                left=Side(style=style.border_style, color=border_color),
                right=Side(style=style.border_style, color=border_color),
                top=Side(style=style.border_style, color=border_color),
                bottom=Side(style=style.border_style, color=border_color)
            )
            cell.border = border
        
        # Alignment
        cell.alignment = Alignment(
            horizontal=style.text_align or "left",
            vertical=style.vertical_align or "top",
            wrap_text=style.text_wrap or False
        )
        
        # Number format
        if style.number_format:
            cell.number_format = style.number_format
    
    def _apply_default_header_style(self, cell, theme: Theme) -> None:
        """Apply default header style."""
        theme_config = self.theme_configs[theme]
        
        cell.font = Font(
            name="Arial",
            size=12,
            bold=True,
            color=theme_config["text_color"]
        )
        
        cell.fill = PatternFill(
            start_color=theme_config["header_color"],
            end_color=theme_config["header_color"],
            fill_type="solid"
        )
        
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        
        cell.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
    
    async def _add_chart(
        self, 
        worksheet, 
        chart_config: ChartConfig, 
        data_source: Dict[str, Any],
        theme: Theme
    ) -> None:
        """Add chart to worksheet."""
        try:
            # Get chart class
            chart_class = self.chart_type_mapping.get(chart_config.chart_type)
            if not chart_class:
                logger.warning(f"Unsupported chart type: {chart_config.chart_type}")
                return
            
            # Create chart
            chart = chart_class()
            chart.title = chart_config.title
            chart.width = chart_config.width / 96  # Convert pixels to inches
            chart.height = chart_config.height / 96  # Convert pixels to inches
            
            # Apply theme colors
            theme_config = self.theme_configs[theme]
            
            # Add data
            if chart_config.data_range:
                # Use data range from worksheet
                data = worksheet[chart_config.data_range]
                chart.add_data(data, titles_from_data=True)
            else:
                # Fetch chart data from visualization service
                chart_data = await self._fetch_chart_data(chart_config, data_source)
                if chart_data:
                    # Add data to worksheet and create chart
                    self._add_chart_data_to_worksheet(worksheet, chart, chart_data)
            
            # Apply style configuration
            if chart_config.style_config:
                self._apply_chart_style(chart, chart_config.style_config, theme)
            
            # Position chart
            position = chart_config.position
            anchor_cell = worksheet.cell(
                row=position.get("row", 1),
                column=position.get("col", 1)
            )
            chart.anchor = anchor_cell.coordinate
            
            # Add chart to worksheet
            worksheet.add_chart(chart)
            
        except Exception as e:
            logger.error(f"Failed to add chart {chart_config.chart_id}: {e}")
    
    async def _fetch_chart_data(
        self, 
        chart_config: ChartConfig, 
        data_source: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Fetch chart data from visualization service."""
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    f"{settings.CHART_SERVICE_URL}/api/v1/charts/generate",
                    json={
                        "chart_id": chart_config.chart_id,
                        "chart_type": chart_config.chart_type,
                        "title": chart_config.title,
                        "data_source": data_source,
                        "format": "data"
                    },
                    timeout=settings.CHART_TIMEOUT_SECONDS
                )
                
                if response.status_code == 200:
                    return response.json()
                else:
                    logger.error(f"Chart service error: {response.status_code}")
                    return None
                    
        except Exception as e:
            logger.error(f"Failed to fetch chart data: {e}")
            return None
    
    def _add_chart_data_to_worksheet(
        self, 
        worksheet, 
        chart, 
        chart_data: Dict[str, Any]
    ) -> None:
        """Add chart data to worksheet and configure chart."""
        # Find empty area for chart data
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Start adding data after existing data
        start_row = max_row + 2
        start_col = max_col + 2
        
        # Add labels
        labels = chart_data.get("labels", [])
        for i, label in enumerate(labels):
            worksheet.cell(row=start_row + i + 1, column=start_col, value=label)
        
        # Add series data
        series_data = chart_data.get("series", [])
        for series_idx, series in enumerate(series_data):
            # Add series name
            worksheet.cell(
                row=start_row, 
                column=start_col + series_idx + 1, 
                value=series.get("name", f"Series {series_idx + 1}")
            )
            
            # Add series values
            values = series.get("values", [])
            for i, value in enumerate(values):
                worksheet.cell(
                    row=start_row + i + 1,
                    column=start_col + series_idx + 1,
                    value=value
                )
        
        # Configure chart data
        if labels and series_data:
            # Add data to chart
            data_range = f"{get_column_letter(start_col + 1)}{start_row}:" \
                        f"{get_column_letter(start_col + len(series_data))}{start_row + len(labels)}"
            chart.add_data(worksheet[data_range], titles_from_data=True)
            
            # Add category axis
            cat_range = f"{get_column_letter(start_col)}{start_row + 1}:" \
                       f"{get_column_letter(start_col)}{start_row + len(labels)}"
            chart.set_categories(worksheet[cat_range])
    
    def _apply_chart_style(
        self, 
        chart, 
        style_config: Dict[str, Any], 
        theme: Theme
    ) -> None:
        """Apply chart style configuration."""
        theme_config = self.theme_configs[theme]
        
        # Apply colors from theme
        if hasattr(chart, 'series') and chart.series:
            colors = [
                theme_config["primary_color"],
                theme_config["secondary_color"],
                theme_config["accent_color"]
            ]
            
            for i, series in enumerate(chart.series):
                color = colors[i % len(colors)]
                if hasattr(series, 'graphicalProperties'):
                    series.graphicalProperties.solidFill = color
    
    def _apply_theme(self, workbook: Workbook, theme: Theme) -> None:
        """Apply theme to workbook."""
        # Theme application is handled at the cell level
        # This method can be extended for workbook-level theme settings
        pass
    
    def _set_workbook_properties(
        self, 
        workbook: Workbook, 
        properties: Dict[str, Any]
    ) -> None:
        """Set workbook properties."""
        props = workbook.properties
        
        if "title" in properties:
            props.title = properties["title"]
        if "subject" in properties:
            props.subject = properties["subject"]
        if "creator" in properties:
            props.creator = properties["creator"]
        if "description" in properties:
            props.description = properties["description"]
        if "keywords" in properties:
            props.keywords = properties["keywords"]
        if "category" in properties:
            props.category = properties["category"]
    
    def _apply_validation_rules(
        self, 
        workbook: Workbook, 
        validation_rules: List[DataValidationRule]
    ) -> None:
        """Apply data validation rules."""
        for rule in validation_rules:
            # Find worksheet (assuming first worksheet for now)
            worksheet = workbook.active
            
            # Create data validation
            dv = DataValidation(
                type=rule.validation_type,
                formula1=rule.formula1,
                formula2=rule.formula2,
                showDropDown=rule.show_dropdown,
                showInputMessage=bool(rule.input_message),
                showErrorMessage=bool(rule.error_message)
            )
            
            # Set messages
            if rule.input_title:
                dv.promptTitle = rule.input_title
            if rule.input_message:
                dv.prompt = rule.input_message
            if rule.error_title:
                dv.errorTitle = rule.error_title
            if rule.error_message:
                dv.error = rule.error_message
            
            # Add to worksheet
            worksheet.add_data_validation(dv)
            dv.add(rule.cell_range)
    
    def _apply_workbook_protection(
        self, 
        workbook: Workbook, 
        protection: Dict[str, Any]
    ) -> None:
        """Apply workbook protection."""
        if protection.get("password"):
            workbook.security = WorkbookProtection(
                workbookPassword=protection["password"],
                lockStructure=protection.get("lock_structure", False),
                lockWindows=protection.get("lock_windows", False)
            )
    
    def _apply_worksheet_protection(
        self, 
        worksheet, 
        protection: Dict[str, Any]
    ) -> None:
        """Apply worksheet protection."""
        if protection.get("password"):
            worksheet.protection = SheetProtection(
                password=protection["password"],
                sheet=protection.get("sheet", False),
                objects=protection.get("objects", False),
                scenarios=protection.get("scenarios", False),
                formatCells=protection.get("format_cells", False),
                formatColumns=protection.get("format_columns", False),
                formatRows=protection.get("format_rows", False),
                insertColumns=protection.get("insert_columns", False),
                insertRows=protection.get("insert_rows", False),
                insertHyperlinks=protection.get("insert_hyperlinks", False),
                deleteColumns=protection.get("delete_columns", False),
                deleteRows=protection.get("delete_rows", False),
                selectLockedCells=protection.get("select_locked_cells", True),
                selectUnlockedCells=protection.get("select_unlocked_cells", True),
                sort=protection.get("sort", False),
                autoFilter=protection.get("auto_filter", False),
                pivotTables=protection.get("pivot_tables", False)
            )
    
    async def _save_workbook(
        self, 
        job_id: int, 
        workbook: Workbook, 
        output_format: ExcelFormat,
        filename: str
    ) -> Dict[str, Any]:
        """Save workbook to file."""
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
        
        if output_format == ExcelFormat.XLSX:
            file_extension = ".xlsx"
        elif output_format == ExcelFormat.XLS:
            file_extension = ".xls"
        elif output_format == ExcelFormat.CSV:
            file_extension = ".csv"
        elif output_format == ExcelFormat.ODS:
            file_extension = ".ods"
        else:
            file_extension = ".xlsx"
        
        filename = f"{safe_filename}_{timestamp}_{job_id}{file_extension}"
        file_path = os.path.join(settings.EXCEL_OUTPUT_DIR, filename)
        
        # Save workbook
        if output_format == ExcelFormat.CSV:
            # Save as CSV (first worksheet only)
            worksheet = workbook.active
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data[1:], columns=data[0] if data else [])
            df.to_csv(file_path, index=False)
        else:
            # Save as Excel format
            workbook.save(file_path)
        
        # Get file info
        file_size = os.path.getsize(file_path)
        
        # Count rows and worksheets
        if output_format == ExcelFormat.CSV:
            row_count = len(data) - 1 if data else 0
            worksheet_count = 1
            chart_count = 0
        else:
            row_count = sum(ws.max_row for ws in workbook.worksheets)
            worksheet_count = len(workbook.worksheets)
            chart_count = sum(len(ws._charts) for ws in workbook.worksheets)
        
        return {
            "file_path": file_path,
            "filename": filename,
            "file_size": file_size,
            "row_count": row_count,
            "worksheet_count": worksheet_count,
            "chart_count": chart_count
        }
    
    async def _update_job_status(self, job_id: int, status: JobStatus) -> None:
        """Update job status."""
        query = """
            UPDATE excel_jobs 
            SET status = $1, started_at = $2 
            WHERE id = $3
        """
        await execute_query(query, status.value, datetime.utcnow(), job_id)
    
    async def _update_job_completion(
        self, 
        job_id: int, 
        file_info: Dict[str, Any],
        generation_time: int
    ) -> None:
        """Update job completion."""
        query = """
            UPDATE excel_jobs 
            SET status = $1, file_path = $2, file_size = $3, 
                row_count = $4, worksheet_count = $5, chart_count = $6,
                generation_time_ms = $7, completed_at = $8
            WHERE id = $9
        """
        await execute_query(
            query,
            JobStatus.COMPLETED.value,
            file_info["file_path"],
            file_info["file_size"],
            file_info["row_count"],
            file_info["worksheet_count"],
            file_info["chart_count"],
            generation_time,
            datetime.utcnow(),
            job_id
        )
    
    async def _update_job_error(
        self, 
        job_id: int, 
        error_message: str,
        generation_time: int
    ) -> None:
        """Update job error."""
        query = """
            UPDATE excel_jobs 
            SET status = $1, error_message = $2, generation_time_ms = $3, completed_at = $4
            WHERE id = $5
        """
        await execute_query(
            query,
            JobStatus.FAILED.value,
            error_message,
            generation_time,
            datetime.utcnow(),
            job_id
        )
    
    async def cancel_job(self, job_id: int) -> bool:
        """Cancel a running job."""
        job_key = f"job:{job_id}"
        
        if job_key in self.active_jobs:
            # Remove from active jobs
            self.active_jobs.pop(job_key, None)
            
            # Update database
            query = """
                UPDATE excel_jobs 
                SET status = $1, completed_at = $2
                WHERE id = $3
            """
            await execute_query(
                query,
                JobStatus.CANCELLED.value,
                datetime.utcnow(),
                job_id
            )
            
            return True
        
        return False
    
    async def get_job_status(self, job_id: int) -> Optional[Dict[str, Any]]:
        """Get job status."""
        # Check if job is in active jobs
        job_key = f"job:{job_id}"
        if job_key in self.active_jobs:
            job_info = self.active_jobs[job_key]
            return {
                "job_id": job_id,
                "status": job_info["status"],
                "runtime_seconds": time.time() - job_info["start_time"]
            }
        
        # Get from database
        query = """
            SELECT id, status, created_at, started_at, completed_at, file_path, error_message
            FROM excel_jobs
            WHERE id = $1
        """
        result = await execute_query(query, job_id)
        
        if result:
            return {
                "id": result["id"],
                "status": result["status"],
                "created_at": result["created_at"],
                "started_at": result["started_at"],
                "completed_at": result["completed_at"],
                "file_path": result["file_path"],
                "error_message": result["error_message"]
            }
        
        return None
    
    async def cleanup_old_files(self, days: int) -> None:
        """Clean up old Excel files."""
        try:
            # Get old jobs
            query = """
                SELECT id, file_path
                FROM excel_jobs
                WHERE created_at < NOW() - INTERVAL '%s days'
                AND file_path IS NOT NULL
            """
            old_jobs = await execute_query(query, days)
            
            if not old_jobs:
                return
            
            # Delete files
            for job in old_jobs:
                file_path = job["file_path"]
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        logger.info(f"Deleted old file: {file_path}")
                    except Exception as e:
                        logger.error(f"Failed to delete file {file_path}: {e}")
            
            # Update database
            job_ids = [job["id"] for job in old_jobs]
            query = """
                UPDATE excel_jobs 
                SET file_path = NULL, file_size = NULL
                WHERE id = ANY($1)
            """
            await execute_query(query, job_ids)
            
            logger.info(f"Cleaned up {len(old_jobs)} old Excel files")
            
        except Exception as e:
            logger.error(f"Cleanup failed: {e}")


# Global instance
excel_generator = ExcelGenerator()