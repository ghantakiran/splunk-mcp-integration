"""
Tests for Excel generator service.
"""

import pytest
import os
from unittest.mock import Mock, patch, AsyncMock
from datetime import datetime

from app.services.excel_generator import ExcelGenerator, ExcelGenerationError
from app.models.excel_models import (
    WorkbookConfig, WorksheetConfig, ChartConfig, CellData, CellStyle,
    ExcelFormat, Theme, ChartType, CellDataType, JobStatus
)


@pytest.fixture
def excel_generator():
    """Create Excel generator instance."""
    return ExcelGenerator()


@pytest.fixture
def mock_workbook_config(sample_worksheet_config):
    """Create mock workbook configuration."""
    return WorkbookConfig(
        name="Test Workbook",
        worksheets=[sample_worksheet_config],
        theme=Theme.OFFICE
    )


class TestExcelGenerator:
    """Test Excel generator functionality."""
    
    @pytest.mark.asyncio
    async def test_theme_configs_loaded(self, excel_generator):
        """Test theme configurations are loaded."""
        assert excel_generator.theme_configs is not None
        assert len(excel_generator.theme_configs) == 5
        assert Theme.OFFICE in excel_generator.theme_configs
        assert Theme.MODERN in excel_generator.theme_configs
        assert Theme.COLORFUL in excel_generator.theme_configs
        assert Theme.DARK in excel_generator.theme_configs
        assert Theme.LIGHT in excel_generator.theme_configs
    
    @pytest.mark.asyncio
    async def test_chart_type_mapping(self, excel_generator):
        """Test chart type mapping is configured."""
        assert excel_generator.chart_type_mapping is not None
        assert len(excel_generator.chart_type_mapping) == 8
        assert ChartType.LINE in excel_generator.chart_type_mapping
        assert ChartType.BAR in excel_generator.chart_type_mapping
        assert ChartType.PIE in excel_generator.chart_type_mapping
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_update_job_status(self, mock_execute_query, excel_generator):
        """Test job status update."""
        mock_execute_query.return_value = None
        
        await excel_generator._update_job_status(1, JobStatus.PROCESSING)
        
        mock_execute_query.assert_called_once()
        args = mock_execute_query.call_args[0]
        assert JobStatus.PROCESSING.value in args
        assert 1 in args
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_update_job_completion(self, mock_execute_query, excel_generator):
        """Test job completion update."""
        mock_execute_query.return_value = None
        
        file_info = {
            'file_path': '/tmp/test.xlsx',
            'file_size': 1024,
            'row_count': 100,
            'worksheet_count': 2,
            'chart_count': 1
        }
        
        await excel_generator._update_job_completion(1, file_info, 5000)
        
        mock_execute_query.assert_called_once()
        args = mock_execute_query.call_args[0]
        assert JobStatus.COMPLETED.value in args
        assert file_info['file_path'] in args
        assert file_info['file_size'] in args
        assert 5000 in args
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_update_job_error(self, mock_execute_query, excel_generator):
        """Test job error update."""
        mock_execute_query.return_value = None
        
        error_message = "Test error message"
        
        await excel_generator._update_job_error(1, error_message, 3000)
        
        mock_execute_query.assert_called_once()
        args = mock_execute_query.call_args[0]
        assert JobStatus.FAILED.value in args
        assert error_message in args
        assert 3000 in args
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_cancel_job(self, mock_execute_query, excel_generator):
        """Test job cancellation."""
        mock_execute_query.return_value = None
        
        # Add job to active jobs
        excel_generator.active_jobs["job:1"] = {
            "job_id": 1,
            "start_time": 1234567890,
            "status": JobStatus.PROCESSING
        }
        
        result = await excel_generator.cancel_job(1)
        
        assert result is True
        assert "job:1" not in excel_generator.active_jobs
        mock_execute_query.assert_called_once()
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_get_job_status_from_database(self, mock_execute_query, excel_generator):
        """Test job status retrieval from database."""
        mock_job = {
            "id": 1,
            "status": "completed",
            "created_at": datetime.now(),
            "file_path": "/tmp/test.xlsx"
        }
        mock_execute_query.return_value = mock_job
        
        status = await excel_generator.get_job_status(1)
        
        assert status is not None
        assert status["id"] == 1
        assert status["status"] == "completed"
        mock_execute_query.assert_called_once()
    
    @pytest.mark.asyncio
    async def test_get_job_status_from_active_jobs(self, excel_generator):
        """Test job status retrieval from active jobs."""
        import time
        
        start_time = time.time()
        excel_generator.active_jobs["job:1"] = {
            "job_id": 1,
            "start_time": start_time,
            "status": JobStatus.PROCESSING
        }
        
        with patch('time.time', return_value=start_time + 10):
            status = await excel_generator.get_job_status(1)
        
        assert status is not None
        assert status["job_id"] == 1
        assert status["status"] == JobStatus.PROCESSING
        assert status["runtime_seconds"] == 10
    
    @pytest.mark.asyncio
    @patch('app.services.excel_generator.execute_query')
    async def test_cleanup_old_files(self, mock_execute_query, excel_generator, temp_dir):
        """Test cleanup of old files."""
        # Create test file
        test_file = os.path.join(temp_dir, "test.xlsx")
        with open(test_file, 'w') as f:
            f.write("test content")
        
        # Mock database response
        mock_execute_query.return_value = [
            {"id": 1, "file_path": test_file}
        ]
        
        await excel_generator.cleanup_old_files(1)
        
        # File should be deleted
        assert not os.path.exists(test_file)
        # Database should be updated
        assert mock_execute_query.call_count == 2
    
    @pytest.mark.asyncio
    async def test_apply_cell_style(self, excel_generator):
        """Test cell style application."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="Test")
        
        style = CellStyle(
            font_name="Arial",
            font_size=12,
            font_bold=True,
            font_color="FF0000",
            background_color="FFFF00",
            border_style="thin",
            text_align="center"
        )
        
        excel_generator._apply_cell_style(cell, style, Theme.OFFICE)
        
        assert cell.font.name == "Arial"
        assert cell.font.size == 12
        assert cell.font.bold is True
        assert cell.alignment.horizontal == "center"
    
    @pytest.mark.asyncio
    async def test_apply_default_header_style(self, excel_generator):
        """Test default header style application."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="Header")
        
        excel_generator._apply_default_header_style(cell, Theme.OFFICE)
        
        assert cell.font.bold is True
        assert cell.font.size == 12
        assert cell.alignment.horizontal == "center"
    
    @pytest.mark.asyncio
    async def test_populate_cell_with_formula(self, excel_generator):
        """Test cell population with formula."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        
        cell_data = CellData(
            value=None,
            formula="SUM(A1:A10)",
            data_type=CellDataType.FORMULA
        )
        
        excel_generator._populate_cell(cell, cell_data, Theme.OFFICE)
        
        assert cell.value == "=SUM(A1:A10)"
    
    @pytest.mark.asyncio
    async def test_populate_cell_with_number(self, excel_generator):
        """Test cell population with number."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        
        cell_data = CellData(
            value=1234.56,
            data_type=CellDataType.NUMBER
        )
        
        excel_generator._populate_cell(cell, cell_data, Theme.OFFICE)
        
        assert cell.value == 1234.56
    
    @pytest.mark.asyncio
    async def test_populate_cell_with_currency(self, excel_generator):
        """Test cell population with currency."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        
        cell_data = CellData(
            value=1234.56,
            data_type=CellDataType.CURRENCY
        )
        
        excel_generator._populate_cell(cell, cell_data, Theme.OFFICE)
        
        assert cell.value == 1234.56
        assert cell.number_format == '"$"#,##0.00'
    
    @pytest.mark.asyncio
    async def test_populate_cell_with_percentage(self, excel_generator):
        """Test cell population with percentage."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        
        cell_data = CellData(
            value=0.25,
            data_type=CellDataType.PERCENTAGE
        )
        
        excel_generator._populate_cell(cell, cell_data, Theme.OFFICE)
        
        assert cell.value == 0.25
        assert cell.number_format == '0.00%'
    
    @pytest.mark.asyncio
    async def test_set_workbook_properties(self, excel_generator):
        """Test workbook properties setting."""
        from openpyxl import Workbook
        
        wb = Workbook()
        properties = {
            "title": "Test Workbook",
            "subject": "Test Subject",
            "creator": "Test Creator",
            "description": "Test Description"
        }
        
        excel_generator._set_workbook_properties(wb, properties)
        
        assert wb.properties.title == "Test Workbook"
        assert wb.properties.subject == "Test Subject"
        assert wb.properties.creator == "Test Creator"
        assert wb.properties.description == "Test Description"
    
    @pytest.mark.asyncio
    @patch('httpx.AsyncClient')
    async def test_fetch_chart_data_success(self, mock_client, excel_generator):
        """Test successful chart data fetching."""
        # Mock HTTP response
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.json.return_value = {
            "labels": ["A", "B", "C"],
            "series": [{"name": "Series 1", "values": [1, 2, 3]}]
        }
        
        mock_client_instance = Mock()
        mock_client_instance.post.return_value = mock_response
        mock_client.return_value.__aenter__.return_value = mock_client_instance
        
        chart_config = ChartConfig(
            chart_id="test_chart",
            chart_type=ChartType.BAR,
            title="Test Chart",
            width=600,
            height=400,
            position={"row": 1, "col": 1}
        )
        
        result = await excel_generator._fetch_chart_data(chart_config, {})
        
        assert result is not None
        assert "labels" in result
        assert "series" in result
    
    @pytest.mark.asyncio
    @patch('httpx.AsyncClient')
    async def test_fetch_chart_data_failure(self, mock_client, excel_generator):
        """Test chart data fetching failure."""
        # Mock HTTP response
        mock_response = Mock()
        mock_response.status_code = 500
        
        mock_client_instance = Mock()
        mock_client_instance.post.return_value = mock_response
        mock_client.return_value.__aenter__.return_value = mock_client_instance
        
        chart_config = ChartConfig(
            chart_id="test_chart",
            chart_type=ChartType.BAR,
            title="Test Chart",
            width=600,
            height=400,
            position={"row": 1, "col": 1}
        )
        
        result = await excel_generator._fetch_chart_data(chart_config, {})
        
        assert result is None
    
    @pytest.mark.asyncio
    async def test_add_chart_data_to_worksheet(self, excel_generator):
        """Test adding chart data to worksheet."""
        from openpyxl import Workbook
        from openpyxl.chart import BarChart
        
        wb = Workbook()
        ws = wb.active
        chart = BarChart()
        
        chart_data = {
            "labels": ["A", "B", "C"],
            "series": [
                {"name": "Series 1", "values": [1, 2, 3]},
                {"name": "Series 2", "values": [4, 5, 6]}
            ]
        }
        
        excel_generator._add_chart_data_to_worksheet(ws, chart, chart_data)
        
        # Check that data was added to worksheet
        assert ws.cell(row=2, column=1).value == "A"
        assert ws.cell(row=3, column=1).value == "B"
        assert ws.cell(row=4, column=1).value == "C"
        assert ws.cell(row=1, column=2).value == "Series 1"
        assert ws.cell(row=2, column=2).value == 1
        assert ws.cell(row=3, column=2).value == 2
        assert ws.cell(row=4, column=2).value == 3